"""
Excel Export Service — converts extracted product data to styled .xlsx files.
"""

import io
from typing import Any

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


def export_to_excel(data: list[dict[str, Any]], filename: str = "extracted_data") -> bytes:
    """
    Convert a list of product dicts to a styled Excel file.

    Returns raw bytes of the .xlsx file ready for HTTP download.
    """
    if not data:
        # Return an empty sheet with a message
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        ws["A1"] = "No data extracted yet."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Normalise data: collect all unique column keys ──
    all_keys: list[str] = []
    for row in data:
        for k in row.keys():
            if k not in all_keys:
                all_keys.append(k)

    # Build DataFrame (fills missing fields with empty string)
    df = pd.DataFrame(data, columns=all_keys).fillna("")

    # ── Write to Excel with openpyxl for styling ──────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Products"

    # Header style
    header_fill   = PatternFill("solid", fgColor="1E293B")  # dark slate
    header_font   = Font(bold=True, color="F8FAFC", size=11)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155"),
    )

    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_prettify(col_name))
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    # Alternating row colours
    even_fill = PatternFill("solid", fgColor="F1F5F9")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(size=10)
    data_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
            cell.fill   = fill
            cell.font   = data_font
            cell.alignment = data_align
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 22

    # Auto-size columns (max 50 chars width)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(_prettify(col_name)),
            *(len(str(v)) for v in df[col_name]),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # Freeze top row
    ws.freeze_panes = "A2"

    # ── Auto-filter ──────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Total Products Extracted:"
    ws_summary["B3"] = len(df)
    ws_summary["A4"] = "Total Columns:"
    ws_summary["B4"] = len(df.columns)
    ws_summary["A5"] = "Columns:"
    for i, col in enumerate(df.columns):
        ws_summary.cell(row=5 + i, column=2, value=_prettify(col))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _prettify(key: str) -> str:
    """Convert snake_case or camelCase to Title Case with spaces."""
    import re
    # Insert space before uppercase letters (camelCase)
    key = re.sub(r"([A-Z])", r" \1", key)
    # Replace underscores/hyphens with space
    key = key.replace("_", " ").replace("-", " ")
    return key.strip().title()
